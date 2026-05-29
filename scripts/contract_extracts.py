import os
import django
from decimal import Decimal

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font

# 🔧 Initialisation Django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "papex.settings.prod")
django.setup()

from api.contracts.models import Contract


def get_service_label(service):
    if not service:
        return ""

    for field in ["name", "title", "label", "display_name"]:
        value = getattr(service, field, None)
        if value:
            return value

    return str(service)


def run():
    print("🔎 Recherche des contrats avec paiement à 0 €...")

    contracts = (
        Contract.objects
        .select_related("client", "service", "created_by")
        .prefetch_related("receipts")
        .order_by("-created_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Paiement zero"

    headers = [
        "Contract ID",
        "Client",
        "Service",
        "Montant initial",
        "Remise (%)",
        "Montant réel",
        "Total payé",
        "Remboursé",
        "Payé net",
        "Solde restant",
        "Créé le",
        "Créé par",
        "Signé",
        "Annulé",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    count = 0
    total_amount_due = Decimal("0.00")
    total_real_amount = Decimal("0.00")
    total_paid = Decimal("0.00")
    total_refunded = Decimal("0.00")
    total_net_paid = Decimal("0.00")
    total_balance_due = Decimal("0.00")

    for contract in contracts:
        if contract.net_paid != Decimal("0.00"):
            continue

        amount_due = contract.amount_due or Decimal("0.00")
        real_amount = contract.real_amount
        amount_paid = contract.amount_paid
        refund_amount = contract.refund_amount or Decimal("0.00")
        net_paid = contract.net_paid
        balance_due = contract.balance_due

        count += 1
        total_amount_due += amount_due
        total_real_amount += real_amount
        total_paid += amount_paid
        total_refunded += refund_amount
        total_net_paid += net_paid
        total_balance_due += balance_due

        ws.append([
            contract.id,
            getattr(contract.client, "full_name", contract.client.pk),
            get_service_label(contract.service),
            float(amount_due),
            float(contract.discount_percent or 0),
            float(real_amount),
            float(amount_paid),
            float(refund_amount),
            float(net_paid),
            float(balance_due),
            contract.created_at.strftime("%Y-%m-%d %H:%M"),
            str(contract.created_by) if contract.created_by else "",
            "Oui" if contract.is_signed else "Non",
            "Oui" if contract.is_cancelled else "Non",
        ])

    ws.append([])
    ws.append(["RAPPORT TOTAL"])
    ws.append(["Nombre de contrats avec paiement à 0 €", count])
    ws.append(["Total montants initiaux", float(total_amount_due)])
    ws.append(["Total montants réels après remise", float(total_real_amount)])
    ws.append(["Total payé brut", float(total_paid)])
    ws.append(["Total remboursé", float(total_refunded)])
    ws.append(["Total payé net", float(total_net_paid)])
    ws.append(["Total restant dû", float(total_balance_due)])

    for row in ws.iter_rows(min_row=ws.max_row - 6, max_row=ws.max_row):
        row[0].font = Font(bold=True)

    filename = f"contracts_paiement_zero_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)

    print(f"✅ {count} contrat(s) avec paiement à 0 € trouvé(s)")
    print(f"💰 Total montant initial : {total_amount_due} €")
    print(f"💰 Total montant réel : {total_real_amount} €")
    print(f"💰 Total payé brut : {total_paid} €")
    print(f"💰 Total remboursé : {total_refunded} €")
    print(f"💰 Total payé net : {total_net_paid} €")
    print(f"💰 Total restant dû : {total_balance_due} €")
    print(f"📄 Fichier généré : {filename}")


if __name__ == "__main__":
    run()